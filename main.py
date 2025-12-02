import sys
import trader
import pandas as pd
from utils_efinance import get_fund_history_ef, get_realtime_rate
import efinance as ef
from openpyxl import load_workbook
from trader import ceboro_trend, combine_today_info, ceboro_suggestion

def backtest_funds(file_path, sheet_name, cash):
    # 打开 workbook
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    sheet = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl", header=1, dtype={"基金代码": str})
    sheet["操作建议"] = sheet["操作建议"].astype(str)
    sheet["追踪ETF/指数"] = sheet["追踪ETF/指数"].astype(str)

    for index, row in sheet.iterrows():
        code = str(row["基金代码"]).strip()
        if not code or code == "nan":
            continue

        etf_code = row["追踪ETF/指数"] if not pd.isna(row["追踪ETF/指数"]) else ""

        etf_info = ef.stock.get_base_info(etf_code)
        etf_name = etf_info.get("股票名称", None)
        volume_source = etf_code if 'ETF' in etf_name else ""

        try:
            fund_info = ef.fund.get_base_info(code)
            fund_name = fund_info.get("基金简称", None)
            ws.cell(row=index + 3, column=2, value=fund_name)
            print(f"✅ {code}: {fund_name}")
        except Exception as e:
            print(f"⚠️ 获取 {code} 基金信息失败: {e}")

        df = get_fund_history_ef(code, 1000, volume_source)
        ceboro_trend(df, trader.NewTrendTaStrategy, False, cash)

        print('-----------------------------------------')

    # 保存 Excel
    wb.save(file_path)

def suggest_funds(file_path, sheet_name, indicators):
    # 打开 workbook
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    sheet = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl", header=1, dtype={"基金代码": str})
    sheet["操作建议"] = sheet["操作建议"].astype(str)
    sheet["追踪ETF/指数"] = sheet["追踪ETF/指数"].astype(str)

    for index, row in sheet.iterrows():
        code = str(row["基金代码"]).strip()
        if not code or code == "nan":
            continue

        etf_code = row["追踪ETF/指数"] if not pd.isna(row["追踪ETF/指数"]) else ""

        etf_info = ef.stock.get_base_info(etf_code)
        etf_name = etf_info.get("股票名称", None)
        volume_source = etf_code if 'ETF' in etf_name else ""

        try:
            fund_info = ef.fund.get_base_info(code)
            fund_name = fund_info.get("基金简称", None)
            ws.cell(row=index + 3, column=2, value=fund_name)
            print(f"✅ {code}: {fund_name}")
        except Exception as e:
            print(f"⚠️ 获取 {code} 基金信息失败: {e}")

        if not etf_code:
            try:
                fund_position = ef.fund.get_invest_position(code)
                etf_code = fund_position.iloc[0].get("股票代码", None)
                ws.cell(row=index + 3, column=3, value=etf_code)
            except Exception as e:
                print(f"⚠️ 获取 {code} 基金持仓信息失败: {e}")

        fund_rate, etf_name = get_realtime_rate(code, etf_code)
        ws.cell(row=index + 3, column=4, value=f'{etf_name}')
        estimate = fund_rate or 0.0
        ws.cell(row=index + 3, column=5, value=f'{estimate/100:.2%}')

        df = get_fund_history_ef(code, 100, volume_source)
        df, forecast_nav = combine_today_info(df, estimate/100)
        action = ceboro_suggestion(df, trader.NewTrendTaStrategy, forecast_nav, estimate / 100, indicators)
        ws.cell(row=index + 3, column=8, value=action)

        print('-----------------------------------------')

    # 保存 Excel
    wb.save(file_path)

def backtest_index(index_code, cash):
    from utils_yfinance import get_usa_stock_yf
    df, _, _ = get_usa_stock_yf(index_code, 'current')

    from trader import ceboro_trend
    ceboro_trend(df, trader.OptimizedTaStrategy, True, cash)

def suggest_index(index_code):
    from utils_yfinance import get_usa_stock_yf
    from trader import ceboro_suggestion
    df, price, estimate = get_usa_stock_yf(index_code, 'current')
    ceboro_suggestion(df, trader.OptimizedTaStrategy, price, estimate, True)

# ----------------- 获取历史数据和当日预估，获得操作建议 -----------------
if __name__ == "__main__":

    while True:
        # use = 'funds' | 'stock' | 'backtest_fund' | 'backtest_index' | 'index'
        use_input = input("""请选择使用功能：
        1. 获取<单个基金>历史数据并回测
        2. 获取<所有基金>历史数据并回测
        3. 获取<单个基金>历史数据并获取操作建议
        4. 获取<所有基金>历史数据并获取操作建议
        5. 获取<指数>历史数据并回测
        6. 获取<指数>历史数据并获取操作建议
        7. 退出程序
        
    🔢 输入选项数字（1-6）：""")

        uses = ['backtest_fund', 'backtest_funds', 'suggest_fund', 'suggest_funds', 'backtest_index', 'suggest_index', 'exit']
        use = uses[int(use_input) - 1]

        file_path = "FundEstimate.xlsx"
        sheet_name = "基金操作"
        cash = 10000

        if use == 'backtest_fund':
            fund_code = input("请输入基金代码：")
            etf_code = input("请输入基金追踪的ETF/指数代码（可留空）：")
            full_log = input("是否输出50条日志并绘图（Y/N）：")
            print(f"开始回测 {fund_code} 基金")
            df = get_fund_history_ef(fund_code, 300, etf_code)
            if df is None or not len(df):
                print(f"⚠️ 获取 {fund_code} 基金历史数据失败")
                sys.exit()
            ceboro_trend(df, trader.NewTrendTaStrategy, full_log == 'Y', cash, full_log == 'Y')

        elif use == 'backtest_funds':
            file_path = "FundEstimate.xlsx"
            print(f"开始回测 {file_path} 文件所有基金")
            backtest_funds(file_path, sheet_name, cash)

        elif use == 'suggest_funds':
            print(f"开始获取 {file_path} 文件所有基金操作建议")
            indicators = input("是否获取常用指标（Y/N）：")
            suggest_funds(file_path, sheet_name, indicators == 'Y')

        elif use == 'backtest_index':
            index_code = input("请输入指数代码：")
            print(f"开始回测 {index_code} 指数")
            backtest_index(index_code, cash)

        elif use == 'suggest_fund':
            fund_code = input("请输入基金代码：")
            try:
                estimate = float(input("请输入基金预估涨跌幅（%）："))
                etf_code = input("请输入基金追踪的ETF/指数代码：")
            except Exception as e:
                print(f"⚠️ 输入的基金预估净值有误: {e}")
                sys.exit()
            print(f"开始获取 {fund_code} 基金操作建议")
            df = get_fund_history_ef(fund_code, 100, etf_code)
            if df is None or not len(df):
                print(f"⚠️ 获取 {fund_code} 基金历史数据失败")
                sys.exit()
            df, forecast_nav = combine_today_info(df, estimate)
            ceboro_suggestion(df, trader.NewTrendTaStrategy, forecast_nav, estimate / 100, True)

        elif use == 'suggest_index':
            index_code = input("⌨️ 请输入指数代码：")
            print(f"开始获取 {index_code} 指数操作建议")
            suggest_index(index_code)

        elif use == 'exit':
            print("退出程序")
            break